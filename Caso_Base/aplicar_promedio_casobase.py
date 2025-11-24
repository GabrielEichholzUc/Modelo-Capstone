#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Aplicar Escenario Promedio Monte Carlo al Caso Base
====================================================

Este script aplica el escenario promedio generado por la simulación Monte Carlo
con bootstrap estacional al archivo Parametros_Finales_Base.xlsx
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import shutil

ARCHIVO_EXCEL = 'Parametros_Finales_Base.xlsx'
ESCENARIO_PROMEDIO = '../Simulacion_MonteCarlo/escenarios_montecarlo/escenario_promedio.xlsx'

def aplicar_escenario_promedio():
    """
    Aplica el escenario promedio Monte Carlo al archivo de parámetros del caso base
    """
    
    print("=" * 70)
    print("APLICAR ESCENARIO PROMEDIO MONTE CARLO AL CASO BASE")
    print("=" * 70)
    
    # 1. Cargar escenario promedio
    print(f"\n📂 Cargando escenario promedio...")
    try:
        df_escenario = pd.read_excel(ESCENARIO_PROMEDIO)
        print(f"  ✓ Escenario promedio cargado")
        print(f"  - Filas: {len(df_escenario)}")
        print(f"  - Temporadas: {sorted(df_escenario['Temporada'].unique())}")
        print(f"  - Afluentes: {sorted(df_escenario['Afluente'].unique())}")
    except FileNotFoundError:
        print(f"  ❌ Error: No se encuentra el archivo '{ESCENARIO_PROMEDIO}'")
        print(f"     Ejecuta primero la simulación Monte Carlo")
        return False
    
    # 2. Crear backup del archivo actual
    print(f"\n💾 Creando backup del archivo actual...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_file = f'Parametros_Finales_Base_backup_{timestamp}.xlsx'
    try:
        shutil.copy2(ARCHIVO_EXCEL, backup_file)
        print(f"  ✓ Backup creado: {backup_file}")
    except Exception as e:
        print(f"  ⚠ No se pudo crear backup: {e}")
    
    # 3. Convertir escenario a formato de parámetros (36 filas: 6 afluentes × 6 temporadas)
    print(f"\n🔄 Convirtiendo escenario a formato de parámetros...")
    
    # Crear lista de filas para el nuevo DataFrame
    filas_qa = []
    
    # Primera fila: encabezado con números de semana
    fila_header = ['t', 'a'] + [float(i) for i in range(1, 49)]
    filas_qa.append(fila_header)
    
    # Para cada temporada
    for t in range(1, 7):
        # Obtener datos de esta temporada del escenario
        df_temp = df_escenario[df_escenario['Temporada'] == t].copy()
        df_temp = df_temp.sort_values('Afluente')
        
        # Para cada afluente (en orden)
        for a_idx, a in enumerate(range(1, 7)):
            fila_afluente = df_temp[df_temp['Afluente'] == a]
            
            if len(fila_afluente) > 0:
                # Extraer valores de semanas (columnas Semana_1, Semana_2, ..., Semana_48)
                semanas_cols = [f'Semana_{w}' for w in range(1, 49)]
                
                # Si las columnas no tienen ese nombre, buscar columnas numéricas
                if semanas_cols[0] not in fila_afluente.columns:
                    # Asumir que las columnas son 1, 2, 3, ..., 48
                    semanas_cols = list(range(1, 49))
                
                valores_semanas = fila_afluente[semanas_cols].values.flatten().tolist()
                
                # Crear fila: [t, a, w1, w2, ..., w48]
                fila = [float(t), float(a)] + valores_semanas
                filas_qa.append(fila)
            else:
                print(f"  ⚠ No se encontraron datos para Temporada={t}, Afluente={a}")
                # Crear fila con ceros
                fila = [float(t), float(a)] + [0.0] * 48
                filas_qa.append(fila)
    
    print(f"  ✓ Formato convertido: {len(filas_qa)-1} filas de datos (6 afluentes × 6 temporadas)")
    
    # 4. Actualizar archivo Excel
    print(f"\n📝 Actualizando archivo {ARCHIVO_EXCEL}...")
    
    try:
        # Cargar workbook
        wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
        
        # Verificar si existe la hoja QA_a,w,t
        if 'QA_a,w,t' not in wb.sheetnames:
            print(f"  ⚠ Hoja 'QA_a,w,t' no existe, creándola...")
            ws = wb.create_sheet('QA_a,w,t')
        else:
            ws = wb['QA_a,w,t']
        
        # Limpiar hoja
        ws.delete_rows(1, ws.max_row)
        
        # Escribir datos
        for fila in filas_qa:
            ws.append(fila)
        
        # Guardar
        wb.save(ARCHIVO_EXCEL)
        print(f"  ✓ Archivo actualizado exitosamente")
        
        # Verificar
        df_verificar = pd.read_excel(ARCHIVO_EXCEL, sheet_name='QA_a,w,t', header=None)
        print(f"  ✓ Verificación: {len(df_verificar)} filas en total")
        print(f"    (1 header + 36 datos = 37 filas esperadas)")
        
    except Exception as e:
        print(f"  ❌ Error actualizando archivo: {e}")
        return False
    
    print(f"\n✅ Escenario promedio Monte Carlo aplicado exitosamente al caso base")
    print(f"   Archivo: {ARCHIVO_EXCEL}")
    print(f"   Backup: {backup_file}")
    
    return True

if __name__ == '__main__':
    aplicar_escenario_promedio()
