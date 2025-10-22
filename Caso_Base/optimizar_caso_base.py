"""
Script principal para optimizar el CASO BASE del modelo de 5 temporadas de la cuenca del Laja
Modelo LP completamente lineal con variables de holgura (deficit, betha)
"""

from caso_base_5temporadas import ModeloLaja
from cargar_datos_5temporadas import cargar_parametros_excel
import time
import os

def verificar_prerequisitos():
    """
    Verifica que los archivos necesarios existan antes de ejecutar
    """
    print("\n" + "="*70)
    print("VERIFICANDO PRE-REQUISITOS")
    print("="*70)
    
    prerequisitos_ok = True
    
    # 1. Verificar archivo Excel
    archivo_excel = "Parametros_Finales_Base.xlsx"
    if os.path.exists(archivo_excel):
        print(f"  ✓ Archivo de parámetros encontrado: {archivo_excel}")
    else:
        print(f"  ✗ ERROR: No se encuentra {archivo_excel}")
        prerequisitos_ok = False
    
    # 2. Verificar que exista la hoja de ve_0_precalculado
    if prerequisitos_ok:
        try:
            import pandas as pd
            from openpyxl import load_workbook
            
            wb = load_workbook(archivo_excel, read_only=True)
            hojas = wb.sheetnames
            
            if 'VE_0_precalculado' in hojas:
                print(f"  ✓ Hoja 'VE_0_precalculado' encontrada en Excel")
                
                # Verificar contenido
                df = pd.read_excel(archivo_excel, sheet_name='VE_0_precalculado')
                if len(df) == 10:  # 2 usos x 5 temporadas = 10 filas
                    print(f"  ✓ Valores precalculados de ve_0 correctos (10 registros)")
                else:
                    print(f"  ⚠️  ADVERTENCIA: Se esperaban 10 registros de ve_0, encontrados {len(df)}")
            else:
                print(f"  ✗ ERROR: Falta hoja 'VE_0_precalculado' en Excel")
                print(f"     → Ejecuta primero: python ejemplo_preprocesamiento.py")
                prerequisitos_ok = False
                
            wb.close()
        except Exception as e:
            print(f"  ⚠️  No se pudo verificar hoja ve_0_precalculado: {e}")
    
    # 3. Verificar módulos importados
    try:
        import gurobipy
        print(f"  ✓ Gurobi disponible (versión {gurobipy.gurobi.version()[0]}.{gurobipy.gurobi.version()[1]})")
    except ImportError:
        print(f"  ✗ ERROR: Gurobi no está instalado")
        prerequisitos_ok = False
    
    print("="*70 + "\n")
    
    if not prerequisitos_ok:
        print("⚠️  ALGUNOS PRE-REQUISITOS NO SE CUMPLEN")
        print("   Revisa los mensajes anteriores antes de continuar.\n")
        respuesta = input("¿Deseas continuar de todas formas? (s/n): ")
        if respuesta.lower() != 's':
            return False
    
    return True


def main():
    """
    Función principal para ejecutar la optimización del caso base (5 temporadas)
    """
    print("\n" + "="*70)
    print(" "*15 + "OPTIMIZACIÓN CASO BASE")
    print(" "*10 + "Cuenca del Laja - 5 Temporadas (LP Lineal)")
    print(" "*10 + "Convenio Hidroeléctricas y Riegos")
    print("="*70 + "\n")
    
    # Verificar pre-requisitos
    if not verificar_prerequisitos():
        print("❌ Optimización cancelada\n")
        return
    
    # 1. Cargar datos desde Excel
    print("="*70)
    print("PASO 1: CARGANDO DATOS DESDE EXCEL")
    print("="*70)
    try:
        parametros = cargar_parametros_excel()
        print("\n✓ Parámetros cargados exitosamente")
    except Exception as e:
        print(f"\n❌ ERROR al cargar parámetros: {e}")
        return
    
    # Verificar que ve_0_precalculado existe
    if 've_0_precalculado' not in parametros:
        print("\n❌ ERROR: No se encontraron valores precalculados de ve_0")
        print("   Ejecuta primero: python ejemplo_preprocesamiento.py")
        return
    
    # 2. Crear instancia del modelo
    print("\n" + "="*70)
    print("PASO 2: INICIALIZANDO MODELO DE OPTIMIZACIÓN")
    print("="*70)
    modelo = ModeloLaja()
    print("✓ Instancia del modelo creada")
    
    # 3. Cargar parámetros en el modelo
    print("\n" + "="*70)
    print("PASO 3: CONFIGURANDO PARÁMETROS DEL MODELO")
    print("="*70)
    modelo.cargar_parametros(parametros)
    print("\n✓ Parámetros configurados en el modelo")
    
    # 4. Construir modelo
    print("\n" + "="*70)
    print("PASO 4: CONSTRUYENDO MODELO MATEMÁTICO")
    print("="*70)
    modelo.construir_modelo()
    
    # Mostrar información del modelo construido
    print("\n📊 INFORMACIÓN DEL MODELO:")
    print(f"  Tipo de modelo:         LP (Linear Programming)")
    print(f"  Variables totales:      {modelo.model.NumVars:,}")
    print(f"  Restricciones:          {modelo.model.NumConstrs:,}")
    print(f"  Variables binarias:     {modelo.model.NumBinVars}")
    print(f"  Temporadas:             {len(modelo.T)}")
    print(f"  Semanas por temporada:  {len(modelo.W)}")
    print(f"  Total semanas:          {len(modelo.T) * len(modelo.W)}")
    print(f"  Centrales:              {len(modelo.I)}")
    print(f"  Canales de riego:       {len(modelo.J)}")
    
    # Mostrar parámetros clave
    print(f"\n🎛️  PARÁMETROS CLAVE:")
    print(f"  V_0 (volumen inicial):  {modelo.V_0:,.0f} hm³")
    print(f"  V_min (vol. mínimo):    {modelo.V_min:,.0f} hm³")
    print(f"  V_MAX (vol. máximo):    {modelo.V_MAX:,.0f} hm³")
    print(f"  psi (penaliz. riego):   {modelo.psi:.2f} GWh/hm³")
    print(f"  phi (penaliz. vol):     {modelo.phi:.2f} GWh/hm³")
    print(f"  Qf (filtración):        {modelo.Qf:.2f} m³/s")
    
    # 5. Optimizar
    print("\n" + "="*70)
    print("PASO 5: EJECUTANDO OPTIMIZACIÓN")
    print("="*70)
    
    # Configuración de optimización
    tiempo_limite = 1800  # 30 minutos (el modelo es LP, debería ser rápido)
    gap = 0.01  # 1% de optimalidad
    
    print(f"\n⚙️  CONFIGURACIÓN DEL SOLVER:")
    print(f"  Tiempo límite:          {tiempo_limite} segundos ({tiempo_limite/60:.0f} minutos)")
    print(f"  Gap de optimalidad:     {gap*100:.1f}%")
    print(f"  Solver:                 Gurobi (Método Simplex para LP)")
    print()
    
    print("🚀 Iniciando optimización...")
    print("="*70)
    
    inicio = time.time()
    modelo.optimizar(time_limit=tiempo_limite, mip_gap=gap)
    tiempo_total = time.time() - inicio
    
    # Verificar estado de la solución
    from gurobipy import GRB
    
    if modelo.model.status == GRB.OPTIMAL:
        print("\n✅ OPTIMIZACIÓN EXITOSA")
        print("="*70)
        print(f"  Estado:                 Óptimo")
        print(f"  Valor objetivo:         {modelo.model.ObjVal:,.2f} GWh")
        print(f"  Tiempo de resolución:   {tiempo_total:.2f} segundos ({tiempo_total/60:.2f} minutos)")
        # MIPGap no está disponible para modelos LP puros
        
    elif modelo.model.status == GRB.TIME_LIMIT:
        print("\n⚠️  TIEMPO LÍMITE ALCANZADO")
        print("="*70)
        print(f"  Estado:                 Tiempo límite")
        print(f"  Mejor solución:         {modelo.model.ObjVal:,.2f} GWh")
        print(f"  Tiempo de resolución:   {tiempo_total:.2f} segundos")
        print("\n  Nota: La solución puede no ser óptima, pero es factible.")
        
    else:
        print(f"\n❌ PROBLEMA EN LA OPTIMIZACIÓN")
        print(f"  Estado del modelo: {modelo.model.status}")
        print("\n  Consulta la documentación de Gurobi para el código de estado.")
        return
    
    # 6. Exportar resultados
    print("\n" + "="*70)
    print("PASO 6: EXPORTANDO RESULTADOS")
    print("="*70)
    
    carpeta_resultados = "resultados_caso_base"
    modelo.exportar_resultados(carpeta_salida=carpeta_resultados)
    
    # 7. Resumen de resultados
    print("\n" + "="*70)
    print("RESUMEN DE RESULTADOS")
    print("="*70)
    
    # Calcular estadísticas de las variables de holgura
    import pandas as pd
    
    # Déficits de riego
    df_riego = pd.read_csv(f"{carpeta_resultados}/riego.csv")
    deficit_total = df_riego['Deficit_m3s'].sum()
    num_deficit = (df_riego['Deficit_m3s'] > 0.01).sum()
    
    print(f"\n📊 ESTADÍSTICAS DE RIEGO:")
    print(f"  Total observaciones:    {len(df_riego)}")
    print(f"  Déficits detectados:    {num_deficit}")
    if num_deficit > 0:
        print(f"  Déficit total:          {deficit_total:,.2f} m³/s acumulado")
        print(f"  Déficit promedio:       {df_riego[df_riego['Deficit_m3s'] > 0.01]['Deficit_m3s'].mean():.2f} m³/s")
        print(f"  Déficit máximo:         {df_riego['Deficit_m3s'].max():.2f} m³/s")
    else:
        print(f"  ✅ No hay déficits de riego")
    
    # Holguras de volumen mínimo
    archivo_betha = f"{carpeta_resultados}/holguras_volumen.csv"
    if os.path.exists(archivo_betha):
        df_betha = pd.read_csv(archivo_betha)
        print(f"\n⚠️  VIOLACIONES DE VOLUMEN MÍNIMO:")
        print(f"  Número de violaciones:  {len(df_betha)}")
        print(f"  Violación total:        {df_betha['Betha_hm3'].sum():,.2f} hm³")
        print(f"  Violación promedio:     {df_betha['Betha_hm3'].mean():.2f} hm³")
        print(f"  Violación máxima:       {df_betha['Betha_hm3'].max():.2f} hm³")
    else:
        print(f"\n✅ VOLUMEN MÍNIMO:")
        print(f"  No hay violaciones del volumen mínimo")
    
    # Volúmenes del lago
    df_vol = pd.read_csv(f"{carpeta_resultados}/volumenes_lago.csv")
    print(f"\n💧 VOLUMEN DEL LAGO:")
    print(f"  Volumen inicial:        {df_vol['Volumen_hm3'].iloc[0]:,.2f} hm³")
    print(f"  Volumen final:          {df_vol['Volumen_hm3'].iloc[-1]:,.2f} hm³")
    print(f"  Volumen mínimo:         {df_vol['Volumen_hm3'].min():,.2f} hm³")
    print(f"  Volumen máximo:         {df_vol['Volumen_hm3'].max():,.2f} hm³")
    print(f"  Volumen promedio:       {df_vol['Volumen_hm3'].mean():,.2f} hm³")
    
    # Energía generada
    df_energia = pd.read_csv(f"{carpeta_resultados}/energia_total.csv")
    energia_total = df_energia['Energia_GWh'].sum()
    print(f"\n⚡ GENERACIÓN ELÉCTRICA:")
    print(f"  Energía total:          {energia_total:,.2f} GWh")
    print(f"  Energía promedio/temp:  {energia_total/len(modelo.T):,.2f} GWh")
    
    # Top 5 centrales por generación
    top5 = df_energia.groupby('Central')['Energia_GWh'].sum().sort_values(ascending=False).head(5)
    print(f"\n  Top 5 centrales:")
    for central, energia in top5.items():
        print(f"    Central {central:2d}: {energia:>10,.2f} GWh ({energia/energia_total*100:5.1f}%)")
    
    # 8. Archivos generados
    print("\n" + "="*70)
    print("ARCHIVOS GENERADOS")
    print("="*70)
    print(f"\n📁 Carpeta: {carpeta_resultados}/\n")
    print("  ✓ generacion.csv           - Caudales de generación por central, semana y temporada")
    print("  ✓ vertimientos.csv         - Vertimientos en cada punto")
    print("  ✓ volumenes_lago.csv       - Evolución del volumen del lago")
    print("  ✓ volumenes_por_uso.csv    - Volúmenes disponibles por uso (riego/generación)")
    print("  ✓ extracciones_por_uso.csv - Caudales extraídos por uso")
    print("  ✓ riego.csv                - Retiros, demandas y déficits de riego")
    if os.path.exists(archivo_betha):
        print("  ⚠️  holguras_volumen.csv     - Violaciones de volumen mínimo (betha)")
    print("  ✓ energia_total.csv        - Energía total generada por central y temporada")
    
    # 9. Resumen final
    print("\n" + "="*70)
    print("OPTIMIZACIÓN COMPLETADA EXITOSAMENTE")
    print("="*70)
    print(f"\n⏱️  Tiempo total de ejecución: {tiempo_total:.2f} segundos ({tiempo_total/60:.2f} minutos)")
    print(f"🎯 Valor función objetivo:    {modelo.model.ObjVal:,.2f} GWh")
    print(f"📊 Temporadas simuladas:       {len(modelo.T)}")
    print(f"📅 Semanas simuladas:          {len(modelo.T) * len(modelo.W)}")
    
    print("\n" + "="*70)
    print("\n✅ Para visualizar los resultados, ejecuta:")
    print("   python visualizar_resultados_5temporadas.py\n")
    print("="*70 + "\n")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Optimización interrumpida por el usuario")
        print("="*70 + "\n")
    except Exception as e:
        print(f"\n\n❌ ERROR CRÍTICO: {e}")
        print("="*70)
        import traceback
        traceback.print_exc()
        print("="*70 + "\n")
