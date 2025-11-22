"""
Script para recalcular los valores f_k correctos usando la función polinomial
de ElToro.txt y actualizar el archivo Excel
"""

import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook

def filt_eltoro(cota):
    """
    Función polinomial de orden 4 para filtraciones en El Toro
    Coeficientes del archivo ElToro.txt
    """
    a0 = -133471.205667
    a1 = 251.668765787
    a2 = -0.112314280288
    a3 = -0.000031180464
    a4 = 0.000000022628942
    
    filt = a0 + (a1 * cota) + (a2 * cota**2) + (a3 * cota**3) + (a4 * cota**4)
    return filt


def cot_eltoro(volumen):
    """
    Función para calcular cota a partir de volumen
    Usa interpolación lineal de tabla empírica del archivo ElToro.txt
    """
    datos = np.array([
        0., 48.28954, 97.47766, 147.26746, 197.35679, 248.04517,
        299.43508, 351.82341, 405.0131, 459.20122, 514.48761, 570.97736,
        628.56274, 687.35149, 747.53804, 808.92531, 871.61054, 935.59635,
        1000.88273, 1067.4697, 1135.25477, 1204.23795, 1274.32464, 1345.60945,
        1417.89268, 1491.07712, 1565.25999, 1640.4439, 1716.42655, 1793.41026,
        1871.19533, 1949.97619, 2029.56105, 2110.14171, 2191.52373, 2273.9068,
        2357.18845, 2441.47116, 2526.65244, 2612.83215, 2700.01554, 2788.19472,
        2877.37496, 2967.75593, 3059.03548, 3151.31608, 3244.69758, 3339.17471,
        3434.65552, 3531.13476, 3628.71226, 3727.4905, 3827.36962, 3928.44686,
        4030.62499, 4133.90402, 4238.58084, 4344.35855, 4451.33437, 4559.61076,
        4668.98805, 4779.66329, 4891.53927, 5004.41367, 5118.38896, 5233.46515,
        5349.83928, 5467.31431, 5585.88761, 5705.66164, 5826.53656
    ])
    
    if volumen <= 0:
        cota = 1300.0
    elif volumen >= datos[-1]:
        cota = 1370.0
    else:
        i = np.searchsorted(datos, volumen)
        d1 = volumen - datos[i-1]
        d2 = datos[i] - datos[i-1]
        dc = d1 / d2
        cota = 1300.0 + i - 2 + dc
    
    cota = np.clip(cota, 1300.0, 1370.0)
    return cota


def recalcular_fk():
    """
    Recalcula los valores f_k usando la función polinomial correcta
    """
    print("\n" + "="*80)
    print("RECALCULANDO f_k CON LA FUNCIÓN POLINOMIAL CORRECTA")
    print("="*80 + "\n")
    
    archivo_excel = "Parametros_Nuevos.xlsx"
    
    # 1. Leer v_k actuales (estos están correctos)
    print("Leyendo v_k desde Excel...")
    df_vk = pd.read_excel(archivo_excel, sheet_name='v_k')
    
    if 'k' in df_vk.columns and 'v' in df_vk.columns:
        v_k_data = df_vk[['k', 'v']].copy()
    elif 'k' in df_vk.columns and 'v_k' in df_vk.columns:
        v_k_data = df_vk[['k', 'v_k']].copy()
        v_k_data.columns = ['k', 'v']
    else:
        v_k_data = df_vk.iloc[:, :2].copy()
        v_k_data.columns = ['k', 'v']
    
    print(f"  ✓ Leídos {len(v_k_data)} volúmenes")
    
    # 2. Leer f_k actuales (incorrectos)
    print("\nLeyendo f_k actuales desde Excel...")
    df_fk_viejo = pd.read_excel(archivo_excel, sheet_name='f_k')
    
    if 'k' in df_fk_viejo.columns and 'f' in df_fk_viejo.columns:
        f_k_viejo_data = df_fk_viejo[['k', 'f']].copy()
    elif 'k' in df_fk_viejo.columns and 'f_k' in df_fk_viejo.columns:
        f_k_viejo_data = df_fk_viejo[['k', 'f_k']].copy()
        f_k_viejo_data.columns = ['k', 'f']
    else:
        f_k_viejo_data = df_fk_viejo.iloc[:, :2].copy()
        f_k_viejo_data.columns = ['k', 'f']
    
    print(f"  ✓ Leídos {len(f_k_viejo_data)} valores de f_k")
    
    # 3. Calcular f_k correctos
    print("\nCalculando f_k correctos desde función polinomial...")
    print("-"*80)
    print(f"{'k':<6} {'v_k [hm³]':<12} {'Cota [msnm]':<14} {'f_k_viejo':<14} {'f_k_nuevo':<14} {'Δf [m³/s]':<12}")
    print("-"*80)
    
    resultados = []
    
    for idx, row in v_k_data.iterrows():
        k = int(row['k'])
        v_k = row['v']
        
        # Calcular cota y filtración
        cota = cot_eltoro(v_k)
        f_k_nuevo = filt_eltoro(cota)
        
        # Obtener f_k viejo
        f_k_viejo = f_k_viejo_data[f_k_viejo_data['k'] == k]['f'].values[0]
        
        diferencia = f_k_nuevo - f_k_viejo
        
        print(f"{k:<6} {v_k:<12.2f} {cota:<14.4f} {f_k_viejo:<14.6f} {f_k_nuevo:<14.6f} {diferencia:<12.6f}")
        
        resultados.append({
            'k': k,
            'v_k': v_k,
            'Cota': cota,
            'f_k_viejo': f_k_viejo,
            'f_k_nuevo': f_k_nuevo,
            'diferencia': diferencia
        })
    
    df_resultados = pd.DataFrame(resultados)
    
    # 4. Estadísticas
    print("\n" + "="*80)
    print("ESTADÍSTICAS DE CORRECCIÓN")
    print("="*80)
    print(f"Diferencia máxima: {df_resultados['diferencia'].abs().max():.6f} m³/s")
    print(f"Diferencia promedio: {df_resultados['diferencia'].abs().mean():.6f} m³/s")
    print(f"Diferencia relativa máxima: {(df_resultados['diferencia'].abs() / df_resultados['f_k_nuevo'] * 100).max():.4f} %")
    
    # 5. Crear DataFrame para Excel
    df_fk_nuevo = pd.DataFrame({
        'k': df_resultados['k'].astype(int),
        'f': df_resultados['f_k_nuevo']
    })
    
    # 6. Actualizar Excel
    print("\n" + "="*80)
    print("ACTUALIZANDO EXCEL")
    print("="*80)
    
    # Cargar workbook
    wb = load_workbook(archivo_excel)
    
    # Verificar si existe la hoja f_k
    if 'f_k' in wb.sheetnames:
        print("  Eliminando hoja 'f_k' antigua...")
        del wb['f_k']
    
    # Crear nueva hoja
    print("  Creando hoja 'f_k' nueva...")
    ws = wb.create_sheet('f_k', index=2)  # Insertar después de Generales
    
    # Escribir headers
    ws['A1'] = 'k'
    ws['B1'] = 'f'
    
    # Escribir datos
    for idx, row in df_fk_nuevo.iterrows():
        ws[f'A{idx+2}'] = int(row['k'])
        ws[f'B{idx+2}'] = float(row['f'])
    
    # Formatear
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Guardar
    wb.save(archivo_excel)
    print(f"  ✓ Excel actualizado: {archivo_excel}")
    
    # 7. Guardar comparación en CSV
    csv_file = "resultados/comparacion_fk_viejo_nuevo.csv"
    df_resultados.to_csv(csv_file, index=False)
    print(f"  ✓ Comparación guardada: {csv_file}")
    
    print("\n" + "="*80)
    print("ACTUALIZACIÓN COMPLETADA")
    print("="*80)
    print("\nAhora ejecuta de nuevo analizar_error_filtraciones.py para ver el error REAL")
    print("de la linealización (sin el sesgo de los f_k incorrectos)")
    print("="*80 + "\n")
    
    return df_resultados, df_fk_nuevo


if __name__ == "__main__":
    df_resultados, df_fk_nuevo = recalcular_fk()
